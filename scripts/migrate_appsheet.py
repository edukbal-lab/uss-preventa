#!/usr/bin/env python3
"""Migración de la BD histórica AppSheet (xlsx) a Supabase.

Lee /Users/eduardokbal/Downloads/base datos appsheet.xlsx
Sube:
  - Hoja Relevamiento → tabla proyectos_historicos (subset de columnas útiles)
  - Hoja Materiales   → tabla materiales_historicos (FK por ID_Relevamiento)

Uso:
  python3 scripts/migrate_appsheet.py [ruta_xlsx]

Requiere: openpyxl, requests (stdlib + pip install openpyxl requests)
"""
import os, sys, json, urllib.request, urllib.error
from datetime import date, datetime
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/Users/eduardokbal/Downloads/base datos appsheet.xlsx"
SUPABASE_URL = "https://lmiaajtuhlcapfyuvqwl.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or "sb_publishable_K2dmKFAX0xZeY-eOsiIv6A_2lNUBRI_"
CHUNK = 500


def norm_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def norm_num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def norm_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = str(v).strip()
    # intenta parsear YYYY-MM-DD HH:MM:SS o similar
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def headers_map(ws):
    """Devuelve dict {header: col_idx (1-based)}."""
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def row_dict(ws, row_idx, hmap):
    return {h: ws.cell(row_idx, c).value for h, c in hmap.items()}


def post_chunked(table, rows):
    """POST rows a Supabase en chunks, con upsert por PK (on_conflict=id)."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict=id"
    headers = {
        "Content-Type": "application/json",
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = len(rows)
    for i in range(0, total, CHUNK):
        batch = rows[i : i + CHUNK]
        body = json.dumps(batch).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                print(f"  {table}: {i+len(batch)}/{total} → HTTP {resp.status}")
        except urllib.error.HTTPError as e:
            print(f"  ERROR {table} chunk {i}: HTTP {e.code} — {e.read()[:400].decode('utf-8','replace')}")
            return False
    return True


def main():
    print(f"Leyendo {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ── Relevamiento → proyectos_historicos ───────────────────────────────
    ws = wb["Relevamiento"]
    hm = headers_map(ws)
    proyectos = []
    valid_ids = set()
    for ri in range(2, ws.max_row + 1):
        r = row_dict(ws, ri, hm)
        pid = norm_str(r.get("ID"))
        if not pid:
            continue
        valid_ids.add(pid)
        proyectos.append({
            "id": pid,
            "fecha": norm_date(r.get("Fecha")),
            "lugar": norm_str(r.get("Lugar")),
            "domicilio": norm_str(r.get("Domicilio")),
            "necesidad": norm_str(r.get("Necesidad del relevamiento")),
            "rubro": norm_str(r.get("Rubro")),
            "solicitud_elementos": norm_str(r.get("Solicitud de elementos")),
            "problematica_cliente": norm_str(r.get("Problematica del cliente")),
            "problematica_resolver": norm_str(r.get("Problematica a resolver")),
            "infraestructura": norm_str(r.get("Infraestructura, dimensiones, acceso a internet")),
            "prioridad": norm_str(r.get("Prioriza costo o servicio")),
            "vendedor": norm_str(r.get("Vendedor")),
            "responsable_lugar": norm_str(r.get("Responsable del lugar")),
            "hecho": norm_str(r.get("Hecho")),
            "venta": norm_str(r.get("Venta")),
            "resultado_comercial": norm_str(r.get("Resultado_comercial")),
            "conclusion": norm_str(r.get("Conclusion")),
            "periferico": norm_str(r.get("Periferico")),
            "funnel_color": norm_str(r.get("Funnin")),
            "costo_total": norm_num(r.get("Costo total")),
            "instalacion": norm_num(r.get("Instalacion")),
            "abono": norm_num(r.get("Abono")),
            "cmf": norm_num(r.get("CMF")),
            "descuento": norm_num(r.get("Descuento")),
            "archivo": norm_str(r.get("Archivo")),
            "imagen_plano": norm_str(r.get("Imagen 6 (plano)")),
        })
    print(f"proyectos_historicos: {len(proyectos)} filas a insertar")
    if proyectos:
        ok = post_chunked("proyectos_historicos", proyectos)
        if not ok:
            sys.exit(1)

    # ── Materiales → materiales_historicos ────────────────────────────────
    ws = wb["Materiales"]
    hm = headers_map(ws)
    materiales = []
    dropped_orphans = 0
    for ri in range(2, ws.max_row + 1):
        r = row_dict(ws, ri, hm)
        mid = norm_str(r.get("ID"))
        proj_id = norm_str(r.get("ID_Relevamiento"))
        if not mid:
            continue
        # Si el proyecto padre no existe, saltamos (evitamos FK violation)
        if proj_id and proj_id not in valid_ids:
            dropped_orphans += 1
            continue
        materiales.append({
            "id": mid,
            "proyecto_id": proj_id,
            "fecha": norm_date(r.get("Fecha")),
            "codigo": norm_str(r.get("Codigo")),
            "detalle": norm_str(r.get("Detalle")),
            "cantidad": norm_num(r.get("Cantidad")),
            "costo": norm_num(r.get("Costos")),
            "iva": norm_num(r.get("IVA")),
            "marca": norm_str(r.get("Marca")),
            "proveedor": norm_str(r.get("Proveedor")),
        })
    print(f"materiales_historicos: {len(materiales)} filas a insertar ({dropped_orphans} huérfanos saltados)")
    if materiales:
        ok = post_chunked("materiales_historicos", materiales)
        if not ok:
            sys.exit(1)

    # ── Mano de obra → mano_obra_historica ────────────────────────────────
    ws = wb["Mano de obra"]
    hm = headers_map(ws)
    mano_obra = []
    dropped_mo = 0
    for ri in range(2, ws.max_row + 1):
        r = row_dict(ws, ri, hm)
        moid = norm_str(r.get("ID"))
        proj_id = norm_str(r.get("ID_Relevamiento"))
        if not moid:
            continue
        if proj_id and proj_id not in valid_ids:
            dropped_mo += 1
            continue
        mano_obra.append({
            "id": moid,
            "proyecto_id": proj_id,
            "nombre": norm_str(r.get("Nombre")),
            "categoria": norm_str(r.get("Categoria")),
            "cantidad_hs": norm_num(r.get("Cantidad hs")),
            "costo_por_hora": norm_num(r.get("Costo por hora")),
            "total": norm_num(r.get("Total")),
            "extras": norm_str(r.get("Extras")),
        })
    print(f"mano_obra_historica: {len(mano_obra)} filas a insertar ({dropped_mo} huérfanos saltados)")
    if mano_obra:
        ok = post_chunked("mano_obra_historica", mano_obra)
        if not ok:
            sys.exit(1)

    print("\n✅ Migración completa.")


if __name__ == "__main__":
    main()
